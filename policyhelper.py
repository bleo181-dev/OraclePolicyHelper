# SCRIPT TO ASSIST IN THE POLICY DECISION-MAKING PROCESS FOR OCI
# BY Leonardo Bombonati (l.bombonati@reply.it)

import os
import re
import sys
import argparse
import openpyxl

def search_by_api(folder):
    api_to_search = input("Enter the API operation to search for (e.g., CreateRun) [CASE SENSITIVE]: ")
    results = 0

    # Scan all files in the folder
    for filename in os.listdir(folder):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder, filename)

            # Open the Excel file
            try:
                workbook = openpyxl.load_workbook(file_path)
            except Exception as e:
                print(f"Error opening the file {filename}: {e}")
                continue

            # Scan all sheets in the file
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Search for the API operation in the current sheet
                for row in sheet.iter_rows(values_only=True):
                    if api_to_search in row:
                        results = 1
                        requested_permission = row[1]  # Assuming the requested_permission is in the second column
                        print(f"\nRequested permission: {requested_permission}")
            workbook.close()

            # API fully covered search
            # Scan all sheets in the file
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Search for the API operation in the current sheet
                for row in sheet.iter_rows(values_only=True):
                    if api_to_search in row:
                        # Scan all sheets in the file
                        for other_sheets in workbook.sheetnames:
                            if other_sheets != "Permissions Required for Each A":
                                other_sheet = workbook[other_sheets]
                                # Search for the required permission in the current sheet
                                for row2 in other_sheet.iter_rows(values_only=True):
                                    if api_to_search in row2[2]:
                                        print(f"\n- API fully covered\nService type: {other_sheets}\nVerb: {row2[0]}\n")
                break
            workbook.close()

            # API partially covered search
            # Scan all sheets in the file
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Search for the API operation in the current sheet
                for row in sheet.iter_rows(values_only=True):
                    if api_to_search in row:
                        # Scan all sheets in the file
                        for other_sheets in workbook.sheetnames:
                            if other_sheets != "Permissions Required for Each A":
                                other_sheet = workbook[other_sheets]
                                # Search for the required permission in the current sheet
                                for row2 in other_sheet.iter_rows(values_only=True):
                                    if api_to_search in row2[3]:
                                        print(f"\n- API partially covered\nService type: {other_sheets}\nVerb: {row2[0]}\n")
                break
            workbook.close()

    return results

def search_by_service_type_and_verb(folder):
    service_type = input("Enter the Service Type to search for (e.g., dataflow-run): ").lower()
    verb = input("Enter the Verb to search for (e.g., manage): ").upper()

    results = 0

    for filename in os.listdir(folder):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder, filename)
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
            except Exception as e:
                print(f"Error opening the file {filename}: {e}")
                continue

            if service_type in workbook.sheetnames:
                sheet = workbook[service_type]
                for row in sheet.iter_rows(values_only=True):
                    if verb in row[0].upper():
                        results = 1
                        print(f"\n\nPermissions: {row[1]}\nAPIs Fully Covered: {row[2]}\nAPIs Partially Covered: {row[3]}\n")
            workbook.close()
    return results

def policy_builder(folder):
    subject = input("Enter the policy subject (e.g., A-Admins): ")
    api_to_search = input("Enter the API operation for which you want to create a policy (e.g., CreateRun) [CASE SENSITIVE]: ")
    results = 0

    # Scan all files in the folder
    for filename in os.listdir(folder):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder, filename)

            # Open the Excel file
            try:
                workbook = openpyxl.load_workbook(file_path)
            except Exception as e:
                print(f"Error opening the file {filename}: {e}")
                continue

            # API fully covered search
            # Scan all sheets in the file
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Search for the API operation in the current sheet
                for row in sheet.iter_rows(values_only=True):
                    if api_to_search in row:
                        # Scan all sheets in the file
                        for other_sheets in workbook.sheetnames:
                            if other_sheets != "Permissions Required for Each A":
                                other_sheet = workbook[other_sheets]
                                # Search for the API in the current sheet
                                for row2 in other_sheet.iter_rows(values_only=True):
                                    if api_to_search in row2[2]:
                                        results = 1
                                        print(f"\nPOLICY\nAllow {subject} to {row2[0].lower()} {other_sheets} in <location> where <conditions>\n")
                break
            workbook.close()

            # API partially covered search
            # Scan all sheets in the file
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Search for the API operation in the current sheet
                for row in sheet.iter_rows(values_only=True):
                    if api_to_search in row:
                        # Scan all sheets in the file
                        for other_sheets in workbook.sheetnames:
                            if other_sheets != "Permissions Required for Each A":
                                other_sheet = workbook[other_sheets]
                                # Search for the API in the current sheet
                                for row2 in other_sheet.iter_rows(values_only=True):
                                    if api_to_search in row2[3]:
                                        results = 1
                                        print(f"\nPOLICY\nAllow {subject} to {row2[0].lower()} {other_sheets} in <location> where <conditions>\n")
                break
            workbook.close()

def policy_to_api(folder, policy_string):
    pattern = r"^Allow (\S+) to (\S+) (\S+).*"
    match = re.match(pattern, policy_string, re.IGNORECASE)
    results = 0

    if match:
        verb = match.group(2).upper()
        service_type = match.group(3)
        
        for filename in os.listdir(folder):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(folder, filename)
                try:
                    workbook = openpyxl.load_workbook(file_path, data_only=True)
                except Exception as e:
                    print(f"Error opening the file {filename}: {e}")
                    continue

                if service_type in workbook.sheetnames:
                    sheet = workbook[service_type]
                    for row in sheet.iter_rows(values_only=True):
                        if verb in row[0].upper():
                            results = 1
                            apis = row[2].split(",")  # Split values by comma
                            for api in apis:
                                print(api.strip())  # Print each value with leading/trailing spaces removed
                workbook.close()
    else:
        print("Invalid policy")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Policy Helper for Oracle Cloud")
    parser.add_argument("-p", "--policy", metavar="POLICY", help="Translate policy to API operation")
    args = parser.parse_args()

    excel_folder = "Excels"

    if args.policy:
        results = policy_to_api(excel_folder, args.policy)  # Call the policy_to_api function if -p argument is provided
    else:
        print("\nWelcome to the Policy Helper for Oracle Cloud!\n\nCurrently, I can only assist with the following Infrastructure Services:")
        for policy_file in os.listdir(excel_folder):
            if os.path.isfile(os.path.join(excel_folder, policy_file)):
                name_without_extension = os.path.splitext(policy_file)[0]
                print(name_without_extension)
        choice = input("\n\nLet's get started...\nDo you want to search by API operation (A), Service Type and Verb (B), or create a policy based on an API operation (C)?\nChoice: ").upper()

        if choice == "A":
            results = search_by_api(excel_folder)
        elif choice == "B":
            results = search_by_service_type_and_verb(excel_folder)
        elif choice == "C":
            results = policy_builder(excel_folder)
        else:
            print("Invalid choice.")

        if results == 0:
            print("\nNo results found. :(\nIf you think this is an error, please check if you have respected case sensitivity for option A or the input for options B or C.\nIf you're still uncertain, personally check in the " + excel_folder + " folder if the OCI service is managed.\nExiting..\n")
        else:
            print("\nGoodbye! :)\nExiting...\n")
